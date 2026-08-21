"""BM4 展開分岐 ↔ pbot dom ケースの対応表を計測する。

1. P進大好きbot「拡張Buchholz OCFに伴う順序数表記」(2018-06-09) の
   T / < / dom / X[Y] を忠実移植し、ブログの計算例で回帰テスト。
2. trio.py の BFS 到達標準形を BM4 分岐 (P0 / T0 / T1 / T2 / ORPH) で分類。
3. z=0 の 2 行断片（Buchholz 記法への翻訳が既知）で
   「BM4 分岐 = dom 終端ケース」を全数照合。
4. 3 行固有の構造主張（T2 なら delta0>0 かつ delta1>0 等）を計測。

結果は ../dom.md の対応表の根拠。
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from trio import parent, expand, diag

# ---------------------------------------------------------------- eB 記法
# 項: 0 | ('D', a, b) = psi_a(b) | ('+', p1, ..., pm) 和 (pi は主要項, m>=2)
ONE = ('D', 0, 0)

def isD(t): return isinstance(t, tuple) and t[0] == 'D'
def isS(t): return isinstance(t, tuple) and t[0] == '+'

def nat(n):
    if n == 0: return 0
    if n == 1: return ONE
    return ('+',) + (ONE,) * n

def as_nat(t):
    if t == 0: return 0
    if t == ONE: return 1
    if isS(t) and all(p == ONE for p in t[1:]): return len(t) - 1
    return None

def tail(t):  # 和の (X2..Xm)
    ps = t[1:]
    return ps[1] if len(ps) == 2 else ('+',) + ps[1:]

def lt(X, Y):
    """ブログ「順序」の忠実移植。"""
    if X == 0: return Y != 0
    if isD(X):
        if Y == 0: return False
        if isD(Y):
            return lt(X[2], Y[2]) if X[1] == Y[1] else lt(X[1], Y[1])
        return X == Y[1] or lt(X, Y[1])
    if Y == 0: return False
    if isD(Y): return lt(X[1], Y)
    if X[1] == Y[1]: return lt(tail(X), tail(Y))
    return lt(X[1], Y[1])

OMEGA = ('D', 0, ONE)

def dom(X):
    if X == 0: return 0
    if isD(X):
        c2 = dom(X[2])
        if c2 == 0:
            c1 = dom(X[1])
            if c1 == 0: return X          # (a-1)
            if c1 == ONE: return X        # (a-2)
            return c1                     # (a-3)
        if c2 == ONE: return OMEGA        # (b)
        if c2 == OMEGA: return OMEGA      # (c)
        if lt(c2, X): return c2           # (d-1)
        return OMEGA                      # (d-2)
    return dom(X[-1])                     # 和

def sub(X, Y):
    """X[Y]。ブログ「基本列」の忠実移植。"""
    if X == 0: return 0
    if isD(X):
        X1, X2 = X[1], X[2]
        c2 = dom(X2)
        if c2 == 0:
            c1 = dom(X1)
            if c1 == 0: return 0                       # (a-1)
            if c1 == ONE: return Y                     # (a-2)
            return ('D', sub(X1, Y), 0)                # (a-3)
        if c2 == ONE:                                  # (b)
            k = as_nat(Y)
            if k is None or k == 0: return 0
            inner = ('D', X1, sub(X2, 0))
            return inner if k == 1 else ('+',) + (inner,) * k
        if c2 == OMEGA: return ('D', X1, sub(X2, Y))   # (c)
        if lt(c2, X): return ('D', X1, sub(X2, Y))     # (d-1)
        Z = c2[1]                                      # (d-2) c2 = <Z,0>
        h = as_nat(Y)
        if h is not None and h >= 1:
            prev = sub(X, nat(h - 1))
            if isD(prev) and prev[1] == X1:            # (d-2-1)
                return ('D', X1, sub(X2, ('D', sub(Z, 0), prev[2])))
        return ('D', X1, sub(X2, ('D', sub(Z, 0), 0))) # (d-2-2)
    ps = X[1:]
    r = sub(ps[-1], Y)
    if r == 0:
        rest = ps[:-1]
        return rest[0] if len(rest) == 1 else ('+',) + rest
    if isD(r): return ('+',) + ps[:-1] + (r,)
    return ('+',) + ps[:-1] + r[1:]

def domchain(X):
    """dom 再帰が通る節のリスト。終端は a1/a2/b/d2/zero。
    各要素は (節ラベル, その節の psi ノード or None)。"""
    if X == 0: return [('zero', None)]
    if isD(X):
        c2 = dom(X[2])
        if c2 == 0:
            c1 = dom(X[1])
            if c1 == 0: return [('a1', X)]
            if c1 == ONE: return [('a2', X)]
            return [('a3', X)] + domchain(X[1])
        if c2 == ONE: return [('b', X)]
        if c2 == OMEGA: return [('c', X)] + domchain(X[2])
        if lt(c2, X): return [('d1', X)] + domchain(X[2])
        return [('d2', X)]
    return [('sum', None)] + domchain(X[-1])

# ---------------------------------------------------------------- 回帰テスト
def regress():
    ok = True
    def chk(name, got, want):
        nonlocal ok
        good = got == want
        ok &= good
        print('  %-28s %s' % (name, 'ok' if good else 'NG got=%r want=%r' % (got, want)))
    Om1 = ('D', nat(1), 0)
    Om2 = ('D', nat(2), 0)
    chk('dom(1) = 1', dom(ONE), ONE)
    chk('dom(omega) = omega', dom(OMEGA), OMEGA)
    chk('dom(Om1) = Om1', dom(Om1), Om1)
    chk('omega[1] = 1  (blog F(1))', sub(OMEGA, nat(1)), ONE)
    chk('omega[3] = 3', sub(OMEGA, nat(3)), nat(3))
    chk('Om1[Y] = Y', sub(Om1, OMEGA), OMEGA)
    X = ('D', 0, Om1)                     # psi_0(Omega_1)
    chk('dom(psi0(Om1)) = omega', dom(X), OMEGA)
    chk('psi0(Om1)[0] = psi0(1)', sub(X, 0), OMEGA)
    chk('psi0(Om1)[2] (blog F(2))', sub(X, nat(2)), ('D', 0, ('D', 0, OMEGA)))
    chk('psi0(2)[2] = w+w (blog)', sub(('D', 0, nat(2)), nat(2)), ('+', OMEGA, OMEGA))
    chk('dom(psi_omega(0)) = omega', dom(('D', OMEGA, 0)), OMEGA)
    chk('dom(Om_Om1) = Om1', dom(('D', Om1, 0)), Om1)
    chk('dom(psi_{w+1}(0)) = self', dom(('D', ('+', OMEGA, ONE), 0)),
        ('D', ('+', OMEGA, ONE), 0))
    chk('dom(psi2(Om1)) = Om1 (d-1)', dom(('D', nat(2), Om1)), Om1)
    chk('dom(psi0(Om2)) = omega (d-2)', dom(('D', 0, Om2)), OMEGA)
    chk('psi0(Om2)[0] = psi0(Om1)', sub(('D', 0, Om2), 0), X)
    chk('psi0(Om2)[1] = psi0(psi1(Om1))',
        sub(('D', 0, Om2), nat(1)), ('D', 0, ('D', nat(1), Om1)))
    return ok

# ---------------------------------------------------------------- BM4 分岐分類
def classify(S):
    x = len(S) - 1
    if all(v == 0 for v in S[x]): return ('P0', None, None)
    t = max(y for y in range(3) if S[x][y] > 0)
    r = parent(S, t, x)
    if r is None: return ('ORPH%d' % t, t, None)
    delta = tuple((S[x][y] - S[r][y]) if y < t else 0 for y in range(3))
    return ('T%d' % t, t, (r, delta))

def bfs(seeds, ns, depth, maxlen=14, cap=60000):
    seen = set()
    frontier = [tuple(S) for S in seeds]
    for S in frontier: seen.add(S)
    for _ in range(depth):
        nxt = []
        for S in frontier:
            if len(S) < 2: continue
            for n in ns:
                T = tuple(expand(list(S), n))
                if len(T) <= maxlen and T not in seen:
                    seen.add(T); nxt.append(T)
                    if len(seen) >= cap: return seen
        frontier = nxt
        if not frontier: break
    return seen

# ---------------------------------------------------------------- 2 行断片の照合
def tr2(S):
    """z 全零の標準形 -> Buchholz 記法（自然数添字 = nat(y)）。"""
    def go(cols):
        if not cols: return 0
        ps = []
        i = 0
        while i < len(cols):
            j = i + 1
            while j < len(cols) and cols[j][0] > cols[i][0]: j += 1
            ps.append(('D', nat(cols[i][1]), go(cols[i+1:j])))
            i = j
        return ps[0] if len(ps) == 1 else ('+',) + tuple(ps)
    return go(list(S))

PRED2 = {'P0': 'a1', 'T0': 'b', 'T1': 'd2', 'ORPH1': 'a2'}

def main():
    print('== 1. pbot 移植の回帰テスト ==')
    assert regress(), 'pbot port regression failed'

    print('\n== 2. BFS 到達集合の分岐分類 ==')
    seeds = [diag(3, v, zcap=1) for v in range(5)] + [diag(3, v, zcap=0) for v in range(5)]
    seen = bfs(seeds, ns=(1, 2, 3), depth=8)
    from collections import Counter
    cnt = Counter(); examples = {}
    orph2_noP1 = orph2_P1 = 0
    zy_bad = 0
    for S in seen:
        if not S: continue
        b = classify(S)
        cnt[b[0]] += 1
        examples.setdefault(b[0], S)
        last = S[-1]
        if last[2] == 1 and last[1] == 0: zy_bad += 1
        # 構造主張
        if b[0] == 'T0':
            assert b[2][1] == (0, 0, 0), (S, b)
        if b[0] == 'T1':
            assert b[2][1][0] > 0 and b[2][1][1] == 0 and b[2][1][2] == 0, (S, b)
        if b[0] == 'T2':
            assert b[2][1][0] > 0 and b[2][1][1] > 0 and b[2][1][2] == 0, (S, b)
        if b[0] == 'ORPH2':
            if parent(list(S), 1, len(S) - 1) is None: orph2_noP1 += 1
            else: orph2_P1 += 1
    print('  到達標準形: %d' % len(seen))
    for k in sorted(cnt):
        print('  %-6s %6d   例: %s' % (k, cnt[k], list(examples[k])[:6]))
    print('  ORPH2 内訳: 行1親なし %d / 行1親あり(全部 z=1) %d' % (orph2_noP1, orph2_P1))
    print('  y=0,z=1 の末尾列: %d 個 (0 のはず)' % zy_bad)
    print('  構造主張 (T0: delta=0 / T1: dx>0,dy=dz=0 / T2: dx>0,dy>0,dz=0): 違反なし')

    print('\n== 3. 2 行断片 (z 全零) で BM4 分岐 = dom 終端ケースを全数照合 ==')
    n2 = 0; bad = []
    posok = {'d2': 0, 'a2': 0}; poschk = {'d2': 0, 'a2': 0}
    for S in seen:
        if not S or any(c[2] != 0 for c in S): continue
        n2 += 1
        b = classify(S)
        chain = domchain(tr2(S))
        term = chain[-1][0]
        want = PRED2.get(b[0])
        if term != want:
            bad.append((S, b[0], term))
            continue
        # 位置照合: d2 の発火ノードの添字 = バッドルートの行1値
        #           かつ潰される基数 = 末尾列の行1値
        if b[0] == 'T1':
            node = chain[-1][1]
            r = b[2][0]
            c2 = dom(node[2])
            poschk['d2'] += 1
            if node[1] == nat(S[r][1]) and c2 == ('D', nat(S[-1][1]), 0):
                posok['d2'] += 1
        if b[0] == 'ORPH1':
            node = chain[-1][1]
            poschk['a2'] += 1
            if node == ('D', nat(S[-1][1]), 0):
                posok['a2'] += 1
    print('  2 行標準形: %d / 不一致: %d' % (n2, len(bad)))
    for S, bb, tt in bad[:5]:
        print('    NG %s %s %s' % (list(S), bb, tt))
    print('  d2 発火ノード = バッドルート(添字照合): %d/%d' % (posok['d2'], poschk['d2']))
    print('  a2 終端ノード = 末尾孤児葉:            %d/%d' % (posok['a2'], poschk['a2']))
    assert not bad

    print('\n== 3.5 孤児 (ORPH) の正例: 到達集合には無いので合成ブロックで ==')
    # 標準形の末尾は常に親を持つ（上の計測で ORPH は 0 件）。
    # 孤児が現れるのは W 階層の graft 対象ブロック（先頭 y>0 / z>0）。
    B1 = [(0, 1, 0)]                       # = Omega_1 の葉
    B2 = [(0, 2, 0), (1, 1, 0)]            # = psi_2(Omega_1): 祖先はいるが全部 y>=1
    for B, name in ((B1, 'B1'), (B2, 'B2')):
        b = classify(B)
        chain = domchain(tr2(B))
        term, node = chain[-1]
        okp = term == 'a2' and node == ('D', nat(B[-1][1]), 0)
        print('  %s = %-24s 分岐 %-6s dom 終端 %s %s' %
              (name, B, b[0], term, 'ok' if okp else 'NG'))
        assert b[0] == 'ORPH1' and okp
    B3 = [(0, 1, 1)]                       # 添字レベルの孤児 (z=1, 行2親なし)
    print('  B3 = %-24s 分岐 %-6s (添字が極限かつ潰されない: Wstar2 graft の対象)'
          % (B3, classify(B3)[0]))
    assert classify(B3)[0] == 'ORPH2'

    print('\n== 4. 3 行固有の実例 ==')
    D1 = diag(3, 1, zcap=1); D2 = diag(3, 2, zcap=1)
    print('  D1 = %s' % D1)
    print('  D1[3] = %s  (2 行対角 = 添字上昇 (a-3))' % expand(D1, 3))
    print('  eB 側 tr(D1) = psi0(psi_omega(0)) の domchain: %s'
          % [c for c, _ in domchain(('D', 0, ('D', OMEGA, 0)))])
    print('  D2 = %s' % D2)
    for n in (1, 2): print('  D2[%d] = %s' % (n, expand(D2, n)))
    psiI = [(0, 0, 0), (1, 1, 1), (2, 1, 1), (3, 1, 0), (2, 0, 0)]
    print('  psi(I) 行列 = %s -> 分岐 %s (t=0 なので (b): 同一コピー)' %
          (psiI, classify(psiI)[0]))
    for n in (1, 2): print('  psi(I)[%d] = %s' % (n, expand(psiI, n)))

    print('\n== 5. BM4-Analysis シート照合（行列 <-> psi ラベルの正解表） ==')
    xlsx = os.path.expanduser('~/proofs/papers/BM4-Analysis-2021.4.27.xlsx')
    if not os.path.exists(xlsx):
        print('  (シートが無いのでスキップ)')
    else:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        label = {}
        for name in wb.sheetnames[1:]:
            for r in wb[name].iter_rows(values_only=True):
                if r and isinstance(r[0], str) and r[0].startswith('('):
                    label.setdefault(r[0], r[1])
        WANT = [
            ('(0,0,0)(1,1,1)', 'psi(W_w)'),
            ('(0,0,0)(1,1,1)(2,1,0)', 'psi(W_w*W)'),
            ('(0,0,0)(1,1,1)(2,1,0)(3,1,0)', 'psi(W_w^W)'),
            ('(0,0,0)(1,1,1)(2,1,0)(3,2,0)', 'psi(W_(w+1))'),
            ('(0,0,0)(1,1,1)(2,1,1)', 'psi(W_(w^2))'),
            ('(0,0,0)(1,1,1)(2,1,1)(3,1,0)', 'psi(W_W)'),
            ('(0,0,0)(1,1,1)(2,1,1)(3,1,0)(2,0,0)', 'psi(I)'),
        ]
        allok = True
        for m, w in WANT:
            got = label.get(m)
            good = got == w
            allok &= good
            print('  %-40s %-14s %s' % (m, got, 'ok' if good else 'NG want ' + w))
        assert allok
        # 罠の記録: (0,0,0)(1,1,1)(2,2,0) は全シートに無い（psi(K) 圏の先）
        assert '(0,0,0)(1,1,1)(2,2,0)' not in label
        print('  (0,0,0)(1,1,1)(2,2,0): 全シートに無し（解析範囲 psi(e(K+1)) の先）')
        # 順序でも確認: S[n] は列辞書式で psi(I) 行列や最終行を追い越す
        def mlt(A, B):  # BM 行列の列辞書式順序（接頭辞は小さい）
            for a, b in zip(A, B):
                if a != b: return a < b
            return len(A) < len(B)
        S = [(0, 0, 0), (1, 1, 1), (2, 2, 0)]
        psiIm = [(0, 0, 0), (1, 1, 1), (2, 1, 1), (3, 1, 0), (2, 0, 0)]
        lastm = [(0, 0, 0), (1, 1, 1), (2, 1, 1), (3, 1, 1), (3, 1, 1), (3, 1, 0), (2, 1, 1)]
        s3 = [tuple(c) for c in expand(S, 3)]
        s4 = [tuple(c) for c in expand(S, 4)]
        assert mlt(psiIm, s3) and mlt(s3, S)
        assert mlt(lastm, s4) and mlt(s4, S)
        print('  psi(I)行列 < S[3] < S,  psi(C(1,0;0))*w 行列 < S[4] < S  (S=(0,0,0)(1,1,1)(2,2,0))')

    print('\nall ok')

if __name__ == '__main__':
    main()
