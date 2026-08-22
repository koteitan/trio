# -*- coding: utf-8 -*-
"""Read Tsskyx's BMS Analysis workbook and write a plain-text dump of it.

The workbook is private, so the dump goes outside the repository (tmp/ is
git-ignored) and nothing extracted from it may be committed.

Layout of a sheet
-----------------
Row 1 names four column bands, one per notation system (Buchholz, Bashicu,
Idealized, Rachel).  Their order and width have already changed once between
revisions of the workbook, so the bands are read off row 1 rather than fixed
here.  A band that only holds '=' in row 2 means "same drawing as the Bashicu
band".

Below that the rows are grouped into blocks, one block per ordinal, in
increasing order.  A block is

    (drawing rows, top first)
    (root row)          <- the drawing's bottom row, holds the band's first column
    (matrix row)        <- optional; holds one Bashicu column per cell

Inside a drawing a node lives in one cell: its band-relative column is its
position in the sequence, and its height above the root row is its depth in
the row-0 forest (equal to the x of the matrix column at the same position).
The cell value is the node's label, i.e. the subscript v of the psi_v that the
node stands for.  A matrix cell is one BM column written as a digit string
with the all-zero tail dropped, so '221' is (2,2,1) and '1' is (1,0,0).

A block is closed by the matrix row, or by a blank row when the sheet has no
matrix rows (the "Up to p0(pw)" sheet is drawings only).

usage: python3 extract.py [xlsx] [outdir]
       defaults: ../papers/tsskyx-bms-analysis.xlsx , tmp/tsskyx
"""
import os
import sys

import openpyxl

KEYS = {'Buchholz': 'B', 'Bashicu': 'S', 'Idealized': 'I', 'Rachel': 'R'}
NCOL = 51                                   # the sheets never go past AY


def bands(ws):
    """[(key, name, first column, last column)] read off the header row."""
    heads = []
    for c in range(1, NCOL + 1):
        v = cell(ws.cell(1, c).value)
        if v:
            heads.append((c, v))
    out = []
    for i, (c, name) in enumerate(heads):
        end = heads[i + 1][0] - 1 if i + 1 < len(heads) else NCOL
        key = KEYS.get(name.split()[0])
        if key:
            out.append((key, name, c, end))
    return out


def cell(v):
    """The cell as a string; Excel hands back 221 as the float 221.0."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return str(v).strip()


def read_rows(ws):
    """[(row number, {column number: value})] over the data rows."""
    out = []
    for r in range(2, ws.max_row + 1):
        d = {}
        for c in range(1, NCOL + 1):
            v = cell(ws.cell(r, c).value)
            if v:
                d[c] = v
        out.append((r, d))
    return out


def split_blocks(rows, root_col):
    """Cut the rows into blocks: a matrix row or a blank row closes one.

    root_col is the first column of the Bashicu band; it is filled exactly
    twice per block, by the root row and by the matrix row under it."""
    blocks, cur = [], []
    prev_has_m = False
    for r, d in rows:
        if not d:                                   # blank separator
            if cur:
                blocks.append(cur)
            cur, prev_has_m = [], False
            continue
        has_m = root_col in d
        cur.append((r, d))
        if has_m and prev_has_m:                    # this row is the matrix row
            blocks.append(cur)
            cur, prev_has_m = [], False
            continue
        prev_has_m = has_m
    if cur:
        blocks.append(cur)
    return blocks


def parse_matrix(cells):
    """['000','111'] -> [(0,0,0),(1,1,1)]; the dropped tail comes back as 0."""
    mat = []
    for s in cells:
        if not s.isdigit():
            return None
        col = [int(ch) for ch in s]
        while len(col) < 3:
            col.append(0)
        mat.append(tuple(col))
    return mat


def parse_band(block, c0, c1):
    """(nodes, matrix) for one band of one block.

    nodes = [(position, height, label)], position counted from the band's
    first column and height counted up from the root row.
    """
    rows = [(r, {c: v for c, v in d.items() if c0 <= c <= c1}) for r, d in block]
    rows = [(r, d) for r, d in rows if d]
    if not rows:
        return [], None
    mat = None
    if len(rows) >= 2 and c0 in rows[-1][1] and c0 in rows[-2][1]:
        last = rows[-1][1]
        mat = parse_matrix([last[c] for c in sorted(last)])
        if mat is not None:
            rows = rows[:-1]
    if not rows:
        return [], mat
    root_row = rows[-1][0]
    nodes = []
    for r, d in rows:
        for c in sorted(d):
            nodes.append((c - c0, root_row - r, d[c]))
    nodes.sort()
    return nodes, mat


def forest(nodes):
    """The nodes as a row-0 forest: a node hangs from the nearest node to its
    left one step lower."""
    kids = {i: [] for i in range(len(nodes))}
    roots, stack = [], {}
    for i, (_, h, _) in enumerate(nodes):
        if h == 0 or h - 1 not in stack:
            roots.append(i)
        else:
            kids[stack[h - 1]].append(i)
        stack[h] = i
        for k in list(stack):
            if k > h:
                del stack[k]
    return roots, kids


def render(nodes):
    """The forest as 'label(child,child)' text."""
    roots, kids = forest(nodes)

    def go(i):
        s = nodes[i][2]
        if kids[i]:
            s += '(' + ','.join(go(j) for j in kids[i]) + ')'
        return s
    return '+'.join(go(i) for i in roots)


def dump(path, outdir):
    wb = openpyxl.load_workbook(path, data_only=False)
    os.makedirs(outdir, exist_ok=True)
    total = 0
    for ws in wb.worksheets:
        bd = bands(ws)
        root = dict((k, c0) for k, _, c0, _ in bd)['S']
        blocks = split_blocks(read_rows(ws), root)
        name = ws.title.replace('/', '_')
        with open(os.path.join(outdir, name + '.txt'), 'w') as f:
            f.write('# %s  (%d blocks)\n' % (ws.title, len(blocks)))
            for n, block in enumerate(blocks, 1):
                f.write('\n[%d] rows %d-%d\n' % (n, block[0][0], block[-1][0]))
                for key, _, c0, c1 in bd:
                    nodes, mat = parse_band(block, c0, c1)
                    if not nodes and mat is None:
                        continue
                    if nodes:
                        f.write('  %s hydra  %s\n' % (key, render(nodes)))
                    if mat:
                        f.write('  %s matrix %s\n'
                                % (key, ''.join('(%d,%d,%d)' % c for c in mat)))
        print('%-22s %4d blocks -> %s' % (ws.title, len(blocks),
                                          os.path.join(outdir, name + '.txt')))
        total += len(blocks)
    print('total %d blocks' % total)


if __name__ == '__main__':
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(os.path.dirname(here))
    xlsx = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        root, '..', 'papers', 'tsskyx-bms-analysis.xlsx')
    outdir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(root, 'tmp', 'tsskyx')
    dump(xlsx, outdir)
